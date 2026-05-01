"""
商品数据导出服务
- Excel (.xlsx) 使用 openpyxl
- CSV 使用标准 csv 模块
"""
import csv
import io
import datetime
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False


HEADER_MAP = {
    "title": "商品标题",
    "platform": "来源平台",
    "price": "价格(¥)",
    "stock": "库存",
    "images": "图片链接",
    "specs": "规格",
    "source_url": "货源链接",
    "platform_product_id": "平台商品ID",
    "status": "采集状态",
}

_EXPORT_HEADERS = ["title", "platform", "price", "stock", "source_url", "platform_product_id", "images", "status"]


def export_csv(products: List[Dict[str, Any]]) -> io.StringIO:
    """导出为 CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    # 写表头
    writer.writerow([HEADER_MAP.get(h, h) for h in _EXPORT_HEADERS])
    for p in products:
        row = []
        for h in _EXPORT_HEADERS:
            val = p.get(h, "")
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            row.append(str(val) if val else "")
        writer.writerow(row)
    output.seek(0)
    return output


def export_xlsx(products: List[Dict[str, Any]]) -> io.BytesIO:
    """导出为 Excel (.xlsx)，带样式"""
    if not _XLSX_AVAILABLE:
        raise RuntimeError("openpyxl 未安装，请 pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "采集商品数据"

    # 表头样式
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell_font = Font(name="微软雅黑", size=10)

    # 写表头
    for col, key in enumerate(_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=HEADER_MAP.get(key, key))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for row_idx, p in enumerate(products, 2):
        for col, key in enumerate(_EXPORT_HEADERS, 1):
            val = p.get(key, "")
            if isinstance(val, list):
                val = "\n".join(str(v) for v in val)
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val else "")
            cell.font = cell_font
            cell.border = thin_border

    # 列宽自适应
    col_widths = [35, 10, 12, 8, 45, 20, 40, 10]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_products(products: List[Dict[str, Any]], fmt: str = "xlsx"):
    """
    统一导出入口
    Args:
        products: 商品数据列表
        fmt: "xlsx" 或 "csv"
    Returns:
        (bytes_or_string, mime_type, filename)
    """
    now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    if fmt == "csv":
        data = export_csv(products)
        filename = f"采集商品_{now}.csv"
        return data, "text/csv; charset=utf-8-sig", filename
    else:
        data = export_xlsx(products)
        filename = f"采集商品_{now}.xlsx"
        return data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename
